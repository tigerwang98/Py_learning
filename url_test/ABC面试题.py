# encoding: utf-8
"""
@project = temp
@file = test07
@author= wanghu
@create_time = 2021/7/16 17:00
"""
import sys
from datetime import datetime
import time
import logging
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from docx import Document
import re
logging.basicConfig(level=logging.INFO)


class ExtractWord():
    def __init__(self):
        self.file_path = r'C:\Users\123\Desktop\ABC笔试题\数据开发岗笔试题\客户提供的商品数据.xlsx'
        self.xiyao_path = r'C:\Users\123\Desktop\ABC笔试题\数据开发岗笔试题\ABC导入模板_西药_中成药.xlsx'
        self.zhongyao_path = r'C:\Users\123\Desktop\ABC笔试题\数据开发岗笔试题\ABC导入模板_中药颗粒_中药饮片.xlsx'
        self.get_file_obj()

    def get_file_obj(self):
        self.file_obj = load_workbook(filename=self.file_path)
        self.xiyao_obj = load_workbook(filename=self.xiyao_path)
        self.zhongyao_obj = load_workbook(filename=self.zhongyao_path)

    @property
    def open_xlsx(self):
        sheet = self.file_obj.active
        for row in sheet.rows:
            name = row[2].value                 # 商品名称
            guige = row[4].value                # 规格
            baozhuang_danwei = row[5].value     # 包装单位
            shengchan_cj = row[20].value        # 生产厂家
            jixing = row[16].value              # 剂型
            pizhun_wenhao = row[21].value       # 批准文号
            tiaoxingma = row[1].value           # 条形码
            tongyongming = row[3].value         # 通用名
            yaopin_type = row[24].value         # 药品类型
            chandi = row[19].value              # 生产地
            zidingyi_type = row[23].value       # 自定义类型

            yield [name, guige, baozhuang_danwei, shengchan_cj, jixing, pizhun_wenhao, tiaoxingma, tongyongming, yaopin_type, chandi, zidingyi_type]

    def handle_data(self, data):
        ret = self.handle_xiyao(data[1], data[4])
        # ret[0]，ret[1],ret[2] 剂量，剂量单位，剂量数量
        if not data[5] or '药典' in data[5]:
            self.handle_zhongyao(data, ret[1])
        else:
            for index, item in enumerate(data):
                if not item:
                    data[index] = ' '
            self.insert_to_xls(True, ['11001(没有找到编码在哪里)', data[-4], data[-3], data[-1], data[0], data[3], data[5], '%s'%data[6], ret[0], ret[1], ret[2], data[4], data[2]])

    def handle_xiyao(self, guige, jixing):
        try:
            jiliang = re.search(r'\d+(\.\d+|\d+)[a-zA-Z]{1,2}', guige).group()
        except AttributeError:
            jiliang = ''
        try:
            jl = re.search(r'(\d+\.\d+)|\d+', jiliang).group()
        except AttributeError:
            jl = ''
        try:
            jl_danwei = re.search(r'[a-zA-Z]+', jiliang).group()
        except AttributeError:
            jl_danwei = ''
        try:
            jizhi_count = re.search(r'\d+', guige.split('*')[-1]).group()
        except AttributeError:
            jizhi_count = ''
        try:
            if '胶囊' in jixing:
                jl_danwei = '粒'
            elif '溶液' in jixing:
                jl_danwei = '支'
        except:
            jl_danwei = ''
        return [jl, jl_danwei, jizhi_count]

    def handle_zhongyao(self, data, jijia_danwei):
        tongyongming = data[-4]
        yaopin_type = data[-3]
        zidingyi_type = data[-1]
        chandi = data[-2]
        tiaoxingma = data[6]
        guige = data[1]
        self.insert_to_xls(False, [11001, tongyongming, yaopin_type, zidingyi_type, chandi, '%s'%tiaoxingma, guige, jijia_danwei])

    def run(self):
        logging.info(datetime.now())
        for index, data in enumerate(self.open_xlsx):
            if index == 0:
                continue
            self.handle_data(data)
        logging.info(datetime.now())

    def insert_to_xls(self, Is_xiyao, data):
        if Is_xiyao:
            sheet = self.xiyao_obj.active
            sheet.append(data)
        else:
            sheet = self.zhongyao_obj.active
            sheet.append(data)

    def save(self):
        self.xiyao_obj.save(self.xiyao_path)
        self.zhongyao_obj.save(self.zhongyao_path)

if __name__ == '__main__':
    t = ExtractWord()
    logging.info('开始了')
    t.run()
    t.save()
    logging.info('结束了！')