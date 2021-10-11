# encoding: utf-8
"""
@project = Py_learing
@file = zhiyiyun
@author= wanghu
@create_time = 2021/10/4 23:30
"""
import json
import random
import re
from redis import Redis
import requests
import time
from datetime import datetime
import pymysql
import logging
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Side,Border
import os
logging.basicConfig(level=logging.INFO)


class ZhiYiYun():
    def __init__(self, date):
        # self.today = str(datetime.now()).split(' ')[0]
        self.date = date
        self.file_path = r'C:\Users\123\Desktop\贵安诊所门诊日志\贵安诊所_%s.xlsx' % self.date
        self.createExcel()
        self.fp = load_workbook(filename=self.file_path)
        self.conn, self.cur, self.redis_conn = self.initDatabase()
        self.userToken = self.getUserToken()
        self.zhensuoID = '8a990d915cc0b3f4015d2b57a0dd04d1'
        self.isFilter = False

    @property
    def get_cur_time(self):
        return str(int(time.time() * 1000))

    @property
    def get_patient_4_update(self):
        url = 'https://patient.zhiyijiankang.com/clinicPatient/syClinicPatient'
        header = {
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent': 'okhttp/4.0.1',
        }
        base_param = {
            'detail': '''
                        {"body":{"code":{},"content":{},"param":{"clientVersion":"%s",
                        "clinicId":"8a990d915cc0b3f4015d2b57a0dd04d1"},"synCode":""},
                        "header":{"v":"1.5.1.1","doctorMainId":"ff808081634bdc1b01634cd801f315cd","clinicId":"8a990d915cc0b3f4015d2b57a0dd04d1",
                        "imei":"866174115530715","currentTime":%s,"userToken":"20211005085737227oha94ox","type":"A0301","currVersion":"V_1.5.1.1"}}
                      '''
        }
        param = {}
        param_list = ['1111111111111', '1559354154002', '1625449982550']
        for index, clientVersion in enumerate(param_list):
            print('正在获取所有病人的第%s页数据... ...' % (index + 1))
            param['detail'] = base_param['detail'] % (clientVersion, int(time.time() * 1000))
            res = requests.post(url=url, data=param, headers=header)
            if res:
                yield res.json()['body']['content']

    @property
    def get_random_temperature(self):
        xiaoshudian = random.choice(range(10))
        temprature = '36.%s' % xiaoshudian
        return temprature

    @property
    def get_random_phone(self):
        list = ['139', '138', '137', '136', '135', '134', '159', '158', '157', '150', '151', '152', '188', '187',
                '182', '183', '184', '178', '130', '131', '132', '156', '155', '186', '185', '176', '133', '153',
                '189', '180', '181', '177']
        str = '0123456789'
        phone = random.choice(list) + "".join(random.choice(str) for i in range(8))
        return phone

    @property
    def generate_idCardNo(self):
        alpha = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
        result = ''
        for i in range(4):
            if i == 3:
                alpha.append('X')
            result += random.choice(alpha)
        return result

    def get_random_idCardNo(self, name, age):
        year = re.findall('\d{1,3}岁', age)
        month = re.findall('\d{1,2}月', age)
        day = re.findall('\d{1,2}天', age)
        if not year:
            year = 0
        if not month:
            month = int(datetime.now().month)
        if not day:
            day = 1
        if isinstance(year, list):
            year = int(year[0].replace('岁', ''))
        if isinstance(month, list):
            month = int(month[0].replace('月', ''))
        if isinstance(day, list):
            day = int(day[0].replace('天', ''))
        delta = int(time.time()) - (year * 365 + month * 30 + day) * 3600 * 24
        try:
            timearray = time.localtime(delta)
            birthday = str(time.strftime('%Y-%m-%d', timearray)).replace('-', '')
        except:
            birthday = self.get_birthday_from_db(name)
        return '511322%s%s' % (birthday, self.generate_idCardNo)

    # 1970年以前没法用时间戳，采用从数据库查birthday的方法
    def get_birthday_from_db(self, name):
        sql = '''select birthday from `zhiyiyun` where patientname="%s"''' % name
        self.cur.execute(sql)
        result = self.cur.fetchone()
        return result[0].replace('-', '')

    def get_phone(self, name):
        sql = '''SELECT phone FROM `zhiyiyun` where patientname="%s" and LENGTH(idcardNo)=18''' % name
        self.cur.execute(sql)
        results = self.cur.fetchall()
        if results:
            for result in results:
                return result[0]
        else:
            return self.get_random_phone

    def get_idcardNo(self, name, age):
        logging.info('正在查询身份证号码... ...')
        select_sql = '''SELECT idcardNo FROM `zhiyiyun` where patientname="%s" and LENGTH(idcardNo)=18''' % (name)
        self.cur.execute(select_sql)
        results = self.cur.fetchall()
        if results:
            for result in results:
                return result[0]
        else:
            logging.info('%s没有找到身份证号!' % name)
            logging.info('正在尝试重新随机生成%s的身份证号' % name)
            return self.get_random_idCardNo(name, age)

    def getUserToken(self):
        url = 'https://usersystem.zhiyijiankang.com/login/login_v1'
        header = {
            'Host': 'usersystem.zhiyijiankang.com',
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent': 'okhttp/4.0.1'
        }
        param = {
            'detail': '''
                    {"body":{"code":{},"content":{},
                    "param":{"device":{"branch":"LIO-AN00","brand":"HUAWEI","clientMac":"","deviceName":"LIO-AN00",
                        "deviceNo":"f227b4fcc74a5dd3","deviceType":"3","ratio":"1600x900","sysVersion":"5.1.1"},
                        "loginName":"13438755360","password":"e10adc3949ba59abbe56e057f20f883e"},"synCode":""},
                    "header":{"v":"1.5.1.1","doctorMainId":"","clinicId":"","imei":"866174115530715","currentTime": %s,
                        "userToken":"","type":"A0301","currVersion":"V_1.5.1.1"}}'''
        }
        param['detail'] = param['detail'] % self.get_cur_time
        res = requests.post(url=url, headers=header, data=param)
        if res:
            token = res.json()['body']['content']['userToken']
            return token
        print('登陆失败!')
        raise Exception

    def request_List(self):
        logging.info('正在获取%s的门诊数据... ...' % self.date)
        cur_time = self.get_cur_time
        base_url = 'http://report.zhiyijiankang.com:5396/outpatiantLog/getOutpatiantLog?date=%s&clinicId=%s&searchDate=%s&searchDateEnd=%s&patientName=&medicinalName=&length=%s&flag=ajax&_=%s'
        header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36',
            'Host': 'report.zhiyijiankang.com:5396',
        }
        result = []
        for i in [0, 50]:
            url = base_url % (cur_time, self.zhensuoID, self.date, self.date, i, cur_time)
            res = requests.get(url=url, headers=header)
            if res:
                result.append(json.dumps(res.json()))
        return result

    def parse_todayList(self, json_data):
        logging.info('正在解析当日的门诊数据... ...')
        today_count = 0
        datas = []
        for patient in json_data['preList']:
            today_count += 1
            name = patient['patient_name']
            sex = patient['sex']
            age = patient['age']
            complaint = patient['patients_complaint'].split('；')[0].replace('，，', ',').replace('，', ',')
            phone = patient['phone'] if patient['phone'] != '-' else self.get_phone(name)
            data = {'name': name, 'sex': sex, 'age': age, 'complaint': complaint, 'phone': phone}
            # print(data)
            data['idcardNo'] = self.get_idcardNo(data['name'], data['age'])
            datas.append(data)
        print('今日门诊量: %s' % today_count)
        return datas

    def calculate_birthday(self, year, month, day):
        year = int(datetime.now().year) - int(year[0].replace('岁', ''))
        month = int(datetime.now().month) - int(month[0].replace('月', ''))
        day = int(datetime.now().day) - int(day[0].replace('天', ''))
        if day < 0:
            month -= 1
            day = 30 - abs(day)
        if month < 0:
            year -= 1
            month = 12 - abs(month)
        if year == 0:       year = int(datetime.now().year)
        if day == 0:        day = int(datetime.now().day)
        if month == 0:      month = int(datetime.now().month)
        return '%s-%s-%s' % (year, month, day)

    def generate_new_data(self, datas):
        if not datas:
            raise Exception
        sheet = self.fp.active
        thead = ['日期', '序号', '是否在发病前14天内有境内外中高风险地区或其他有病例报告社区的旅行史和居住史',
                 '患者姓名', '性别', '年龄', '体温(℃)', '主要症状', '流行病学史(有或无)', '患者流向本地或外地',
                 '身份证号', '联系电话']
        sheet.append(thead)
        count = 0
        for data in datas:
            count += 1
            temprature = self.get_random_temperature
            item = [self.date, count, '否', data['name'], data['sex'], data['age'], temprature, data['complaint'], '无', '本地', data['idcardNo'], data['phone']]
            sheet.append(item)
            logging.info(item)

    def filter_data(self, item):
        filter_id = item['id']
        if self.redis_conn.sadd('zhiyiyun', filter_id):
            return 1
        else:
            return 0

    def update_database(self):
        logging.info('正在更新数据库... ...')
        for items in self.get_patient_4_update:
            for item in items:
                if self.filter_data(item):
                    patientID = item['id']
                    homeAddress = item['homeAddress']
                    patientName = item['userName']
                    birthday = item['birthday']
                    sex = '男' if item['sex'] == 1 else '女'
                    phone = item['phone']
                    idcardNo = item['idCardNo']
                    data = {'patientID': patientID, 'homeAddress': homeAddress,
                            'patientName': patientName, 'birthday': birthday,
                            'sex': sex, 'phone': phone, 'idcardNo': idcardNo,}
                    self.insert_data(data)
                else:
                    logging.info('%s:数据库今日更新有重复:%s' % (self.date, item['userName']))

    def insert_data(self, data):
        logging.info('%s正在插入数据库... ... ' % data['patientName'])
        sql = '''INSERT INTO `zhiyiyun`(patientname, nameID, sex, birthday, phone, address, idcardNo) VALUES ("%s", "%s", "%s", "%s", "%s", "%s", "%s")'''
        insert_sql = sql % (data['patientName'], data['patientID'], data['sex'], data['birthday'], data['phone'], data['homeAddress'], data['idcardNo'])
        self.cur.execute(insert_sql)
        self.conn.commit()
        logging.info('%s 插入成功!' % data['patientName'])

    def initDatabase(self):
        conn = pymysql.connect(host='localhost', user='root', password='123456', db='test', charset='utf8')
        redis_conn = Redis(host='localhost', password='', db='7', port=6379)
        cursor = conn.cursor()
        return conn, cursor, redis_conn

    def createExcel(self):
        if os.path.exists(self.file_path):
           os.remove(self.file_path)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "贵安诊所_%s.xlsx" % self.date
        workbook.save(filename=self.file_path)

    def save(self):
        sheet = self.fp.active
        # 调整表格大小
        sheet.column_dimensions['A'].width = 12.38
        sheet.column_dimensions['B'].width = 4
        sheet.column_dimensions['C'].width = 13.25
        sheet.column_dimensions['D'].width = 8.38
        sheet.column_dimensions['E'].width = 4.75
        sheet.column_dimensions['F'].width = 17
        sheet.column_dimensions['G'].width = 7.38
        sheet.column_dimensions['H'].width = 18.5
        sheet.column_dimensions['I'].width = 5.88
        sheet.column_dimensions['J'].width = 6.13
        sheet.column_dimensions['K'].width = 20.5
        sheet.column_dimensions['L'].width = 11.88
        for row in sheet.rows:
            for cell in row:
                alignment =Alignment(horizontal="center", vertical="center", wrap_text=True)
                side = Side(border_style='thin', color='FF000000')
                border = Border(left=side, right=side, top=side, bottom=side)
                cell.alignment = alignment
                cell.border = border
        sheet.set_printer_settings(paper_size=9, orientation='landscape')        # 0是横向

    def exit(self):
        self.fp.save(self.file_path)
        self.fp.close()
        self.redis_conn.close()
        self.cur.close()
        self.conn.close()

    def run(self):
        datas = []
        # 每天需要先更新数据库
        self.update_database()
        page_list = self.request_List()
        for todayList in page_list:
            try:
                infos = json.loads(todayList)['content'][0]
                data = self.parse_todayList(infos)
                datas.extend(data)
            except IndexError:
                logging.info('出错了!可能是今天人数不够50人,请检查原因!')
        self.generate_new_data(datas)
        self.save()
        self.exit()
        # os.startfile(self.file_path, 'print')

def generate_date():
    for month in range(6, 10):
        for day in range(1, 32):
            if month in [6, 9] and day == 31:
                continue
            if day < 10:
                day = '0' + str(day)
            date = '2021-0%s-%s' % (month, day)
            yield date


if __name__ == '__main__':
    # for date in  generate_date():
        date = '2021-10-11'
        t = ZhiYiYun(date)
        t.run()